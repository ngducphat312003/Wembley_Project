from MainUI1 import Ui_MainWindow
from PyQt6.QtCore import  QDate, QDateTime, QTimer
from PyQt6.QtWidgets import QTableWidgetItem, QAbstractItemView
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from page_Initial import PageInitialBackend

class HistoryBackend():
    def __init__(self, main: Ui_MainWindow, conn, KeyPrimary, initial: PageInitialBackend):
        self.parent = main
        self.conn = conn
        self.KeyPrimary = KeyPrimary
        self.initial = initial

        self.parent.btnDefault.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault", "#A4A4A4", "15px")
        self.parent.btnDefault_2.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault_2", "#A4A4A4", "15px")

        self.rowid = []
        self.count_error = 0
        self.SetupStyleTable(self.parent.tbManufacturingHistory)
        self.SetupStyleTable(self.parent.tbErrorProduct)
        self.SetStylePopup(self.parent.dtDayStartHistory)
        self.SetStylePopup(self.parent.dtDayFinishHistory)
        self.SetStylePopup(self.parent.dtDayStartHistory1)
        self.SetStylePopup(self.parent.dtDayFinishHistory1)
        self.CollectingRow()
        self.UpdateManufactoringHistory1Month(self.rowid)
        self.UpdateHistoryTimer = QTimer()
        self.UpdateHistoryTimer.timeout.connect(self.UpdateManufactoringHistory)
        self.UpdateHistoryTimer.start(5000)
        self.UpdateErrorTimer = QTimer()
        self.UpdateErrorTimer.timeout.connect(self.UpdateErrorHistoryTimer)
        self.UpdateErrorTimer.start(5000)
        self.Screening_Current_Date()
        
        self.parent.dtDayStartHistory.dateChanged.connect(self.lock_date_range)    
        self.parent.dtDayFinishHistory.dateChanged.connect(self.lock_date_range)
        self.parent.dtDayStartHistory1.dateChanged.connect(self.lock_date_range)
        self.parent.dtDayFinishHistory1.dateChanged.connect(self.lock_date_range)

        #Hiển thị lịch sử lỗi sản phẩm
        self.UpdateErrorHistory()

        self.parent.btnFilter1.clicked.connect(self.FilterHistory)
        self.parent.btnFilter2.clicked.connect(self.FilterHistory1)
        self.parent.btnDefault.clicked.connect(self.ResetHistory)
        self.parent.btnDefault_2.clicked.connect(self.ResetHistory1)
        self.parent.btnExport.clicked.connect(self.ExportHistory)
        self.parent.btnExport_2.clicked.connect(self.ExportErrorHistory)

        table = self.parent.tbManufacturingHistory
        table.setColumnWidth(0, 100)
        table.setColumnWidth(1, 240)
        table.setColumnWidth(2, 240)
        table.setColumnWidth(3, 240)
        table.setColumnWidth(4, 240)
        table.setColumnWidth(5, 240)
        table.setColumnWidth(6, 240)
        table.setColumnWidth(7, 240)
        table.setColumnWidth(8, 240)

        table1 = self.parent.tbErrorProduct
        table1.setColumnWidth(0, 100)  # Cột STT
        table1.setColumnWidth(1, 240)  # Cột rowid
        table1.setColumnWidth(2, 240)  # Cột rowid
        table1.setColumnWidth(3, 240)  # Cột rowid
        table1.setColumnWidth(4, 240)  # Cột rowid
        table1.setColumnWidth(5, 305)  # Cột rowid

    def SetStyleSheetForbtn(self, btn, background_color,pixel):
        #Style cho nút 
        button = getattr(self.parent, btn)
        button.setStyleSheet(f"""
                QPushButton#{btn} {{
                    border-radius: {pixel};
                    border: 1px solid white;
                    background-color: {background_color};  /* Màu nền mới */
                    color: white;
                    text-align: center;
                    font-family: Inter, sans-serif;
                }}

                QPushButton#{btn}:hover {{
                    background-color: #FFA500;  /* Màu nền khi hover */
                }}

                QPushButton#{btn}:pressed {{
                    padding-left: 5px;
                    padding-top: 5px;
                }}
                """)
    
    def CollectingRow(self):
        self.rowid = []
        self.current_date = QDate.currentDate()
        one_month_ago = self.current_date.addDays(-30)

        # Chuyển sang chuỗi để dùng trong truy vấn SQL
        current_date_str = self.current_date.toString("yyyy-MM-dd")
        one_month_ago_str = one_month_ago.toString("yyyy-MM-dd")

        c = self.conn.cursor()
        query = """
            SELECT rowid, * FROM ProductHistory
            WHERE DATE(StartTime) BETWEEN ? AND ?
        """
        c.execute(query, (one_month_ago_str, current_date_str))
        rows = c.fetchall()

        for row in rows:
            self.rowid.append(row[0])  # rowid là cột tự động của SQLite
    
    def UpdateManufactoringHistory1Month(self,rowid):
        if not rowid:
            return

        c = self.conn.cursor()
        placeholders = ",".join("?" for _ in rowid)  # tạo ?,?,?... theo số lượng rowid
        query = f"SELECT * FROM ProductHistory WHERE rowid IN ({placeholders})"
        
        c.execute(query, rowid)
        rows = c.fetchall()

        self.DisplayOnTable(rows)
    
    def DisplayOnTable(self, rows):
        table = self.parent.tbManufacturingHistory
        current_row_count = table.rowCount()  # Lấy số dòng hiện tại của bảng
        
        # Ẩn row headers (nếu không muốn hiển thị cột đếm hàng bên trái)
        table.verticalHeader().setVisible(False)
        rows = rows[::-1]
        # Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(rows):
            table.insertRow(current_row_count + row_idx)  # Thêm hàng mới vào cuối bảng
            
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 8:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

        # Tự động giãn cột cho vừa nội dung
        table.resizeColumnsToContents()
        table.resizeRowsToContents()
    
    def SetupStyleTable(self, table):
        scrollbar_v = self.scrollbar("vertical", "ffffff", "0055FF")
        scrollbar_h = self.scrollbar("horizontal", "ffffff", "0055FF")
        table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        table.setShowGrid(False)
        style = """
                QTableWidget {
                    background-color: #ffffff;  /* Màu nền bảng */
                    border-radius: 0px;  /* Bo tròn góc */
                    color: #0055FF;  /* Màu chữ */
                    font-family: Inter, sans-serif;
                    font-size: 14px;
                    font-weight: bold;
                }
                QTableWidget::item {
                    border: 1px solid #0055FF;  /* Viền đỏ cho tất cả ô */
                    font-family: Inter, sans-serif;
                    font-size: 14px;
                    font-weight: bold;
                }
                QHeaderView::section {
                background-color: #ffffff;
                border: 1px solid #0055FF;  /* Ẩn viền */
                font-weight: bold;
                font-family: Inter, sans-serif;
                font-size: 16px;
                }
                """
        full_style = style + scrollbar_v + scrollbar_h

        table.setStyleSheet(full_style)
        
    def Screening_Current_Date(self):
        self.current_date = QDate.currentDate()
        one_week_ago = self.current_date.addDays(-6)
        #Hiển thị trên QDateTimeEdit
        self.parent.dtDayStartHistory.setDate(one_week_ago)
        self.parent.dtDayFinishHistory.setDate(self.current_date)
        self.parent.dtDayStartHistory1.setDate(one_week_ago)
        self.parent.dtDayFinishHistory1.setDate(self.current_date)
    
    def SetStylePopup(self, datetime_edit):
        datetime_edit.setCalendarPopup(True)
        calendar = datetime_edit.calendarWidget()
        calendar.setStyleSheet("""
            QCalendarWidget QWidget { 
                background-color: #FFFFFF;
                border: None; 
                border-radius: 0px;  
            }

            QCalendarWidget QToolButton {
                background-color: white;
                color: #1b731b;
                font: bold 14px;
                border: none;
                margin: 5px;
                padding: 5px;
                border-radius: 0px;
            }

            QCalendarWidget QMenu {
                background-color: white;
                color: black;
                border-radius: 0px;
            }

            QCalendarWidget QSpinBox { 
                width: 60px; 
                font-size: 14px;
                background-color: black; 
            }

            QCalendarWidget QAbstractItemView {
                font-size: 14px;
                background-color: white;
                selection-background-color: #1b731b;
                selection-color: white;
                gridline-color: #FFFFFF;
            }

            QCalendarWidget QAbstractItemView:enabled {
                color: #333;
            }

            QCalendarWidget QAbstractItemView:disabled {
                color: #999;
            }
        """)
    
    def scrollbar(self, direction, base_color, color):
        style = f"""
        QScrollBar:{direction} {{
            border: none; 
            background-color: #{base_color}; 
            {'width' if direction == 'vertical' else 'height'}: 10px; 
            margin: 0px 0px 0px 0px; 
        }}
        
        QScrollBar::handle:{direction} {{
            background-color: #{color}; 
            border-radius: 5px; 
        }}
        
        QScrollBar::handle:{direction}:hover {{
            background-color: #4D9EFF; 
        }}
        
        QScrollBar::add-line:{direction}, QScrollBar::sub-line:{direction} {{
            border: none; 
            background: none; 
        }}
        
        QScrollBar::add-page:{direction}, QScrollBar::sub-page:{direction} {{
            background: none;
        }}
        """
        return style
    #Mỗi 5 giây cập nhật dữ liệu ở hàng đầu tiên trong bảng tương ứng với hàng cuối cùng trong database
    def UpdateManufactoringHistory(self):
        ## Lấy dữ liệu từ hàng cuối cùng trong database
        c = self.conn.cursor()
        c.execute("SELECT * FROM ProductHistory ORDER BY rowid DESC LIMIT 1")
        result = c.fetchone()
        if result:
            # print("result", result)
            for col_idx, value in enumerate(result):
                if col_idx < 8:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.parent.tbManufacturingHistory.setItem(0, col_idx + 1, item) 

    def lock_date_range(self):
        # Lấy ngày hiện tại
        current_date = QDate.currentDate()
        # Thực hiện truy vấn SQL để lấy ngày muộn nhất trong StartTime
        c = self.conn.cursor()
        query = """
            SELECT rowid, StartTime FROM ProductHistory
            ORDER BY StartTime ASC LIMIT 1
        """
        c.execute(query)
        result = c.fetchone()
        # print("result", result)
        if result:
            # Lấy ngày muộn nhất từ StartTime (cột thứ 2 trong kết quả)
            latest_start_date = QDate.fromString(result[1].split()[0], "yyyy-MM-dd")

        # Khóa dtDayStartHistory (không cho phép chọn từ ngày hôm nay trở đi)
        self.parent.dtDayStartHistory.setDateRange(latest_start_date, current_date)
        start_date = self.parent.dtDayStartHistory.date()
        # Khóa dtDayFinishHistory (không cho phép chọn từ ngày hôm sau trở đi)
        self.parent.dtDayFinishHistory.setDateRange(start_date, current_date)

        self.parent.dtDayStartHistory1.setDateRange(latest_start_date, current_date)
        start_date1 = self.parent.dtDayStartHistory1.date()
        # Khóa dtDayFinishHistory (không cho phép chọn từ ngày hôm sau trở đi)
        self.parent.dtDayFinishHistory1.setDateRange(start_date1, current_date)
            
    def UpdateErrorHistory(self):
        current_date = QDate.currentDate()
        #Lấy ngày tiếp theo của current_date
        # next_date = current_date.addDays(1)
        #Lọc Timestamp có trong ErrorHistory chứa ngày current_date
        c = self.conn.cursor()
        query = """
            SELECT * FROM ErrorHistory
            WHERE DATE(Timestamp) = ?
        """
        c.execute(query, (current_date.toString("yyyy-MM-dd"),))
        products = c.fetchall()

        table = self.parent.tbErrorProduct
        current_row_count = table.rowCount()  # Lấy số dòng hiện tại của bảng
        
        # Ẩn row headers (nếu không muốn hiển thị cột đếm hàng bên trái)
        table.verticalHeader().setVisible(False)
        rows = products[::-1]
        self.count_error = len(rows)
        # Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(rows):
            table.insertRow(current_row_count + row_idx)  # Thêm hàng mới vào cuối bảng
            
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 5:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự
        
    def UpdateErrorHistoryTimer(self):
        current_date = QDate.currentDate()
        #Lọc Timestamp có trong ErrorHistory chứa ngày current_date
        c = self.conn.cursor()
        query = """
            SELECT * FROM ErrorHistory
            WHERE DATE(Timestamp) = ?
        """
        c.execute(query, (current_date.toString("yyyy-MM-dd"),))
        products = c.fetchall()
        rows = products[::-1]
        if len(rows) != self.count_error:
            table = self.parent.tbErrorProduct
            table.setRowCount(0)
            current_row_count = table.rowCount()
            for row_idx, row_data in enumerate(rows):
                table.insertRow(current_row_count + row_idx)  # Thêm hàng mới vào cuối bảng
                # Thêm cột số thứ tự vào cột đầu tiên (STT)
                item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
                # Thêm các giá trị khác vào các cột còn lại
                for col_idx, value in enumerate(row_data):
                    if col_idx < 5:  # Tránh việc truy cập ngoài phạm vi
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

    def FilterHistory(self):
        self.Push_working_history("Lọc lịch sử sản xuất")
        self.parent.btnDefault.setEnabled(True)
        self.SetStyleSheetForbtn("btnDefault", "#33D909", "15px")
        
        self.UpdateHistoryTimer.stop()
        start_date = self.parent.dtDayStartHistory.date()
        end_date = self.parent.dtDayFinishHistory.date()

        c = self.conn.cursor()
        query = """
            SELECT * FROM ProductHistory
            WHERE DATE(StartTime) BETWEEN ? AND ?
        """
        c.execute(query, (start_date.toString("yyyy-MM-dd"), end_date.toString("yyyy-MM-dd")))
        products = c.fetchall()
        table = self.parent.tbManufacturingHistory
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng

        rows = products[::-1]
        # Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(rows):
            table.insertRow(row_idx)  # Thêm hàng mới vào cuối bảng
            
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 8:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

    def ResetHistory(self):
        self.Push_working_history("Trở về lịch sử sản xuất hiện tại")
        # Đặt lại các giá trị về mặc định
        table = self.parent.tbManufacturingHistory
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
        self.parent.btnDefault.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault", "#A4A4A4", "15px")
        self.UpdateHistoryTimer.start(5000)
        self.CollectingRow()
        self.UpdateManufactoringHistory1Month(self.rowid)

        # Đặt lại giá trị của các QDateTimeEdit về ngày hiện tại
        self.current_date = QDate.currentDate()
        one_week_ago = self.current_date.addDays(-6)
        self.parent.dtDayStartHistory.setDate(one_week_ago)
        self.parent.dtDayFinishHistory.setDate(self.current_date)

    def FilterHistory1(self):
        self.Push_working_history("Lọc lịch sử lỗi sản phẩm")
        self.parent.btnDefault_2.setEnabled(True)
        self.SetStyleSheetForbtn("btnDefault_2", "#33D909", "15px")
        
        self.UpdateErrorTimer.stop()
        start_date = self.parent.dtDayStartHistory1.date()
        end_date = self.parent.dtDayFinishHistory1.date()

        c = self.conn.cursor()
        query = """
            SELECT * FROM ErrorHistory
            WHERE DATE(Timestamp) BETWEEN ? AND ?
        """
        c.execute(query, (start_date.toString("yyyy-MM-dd"), end_date.toString("yyyy-MM-dd")))
        products = c.fetchall()
        table = self.parent.tbErrorProduct
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
        rows = products[::-1]
        # Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(rows):
            table.insertRow(row_idx)  # Thêm hàng mới vào cuối bảng
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 5:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

    def ResetHistory1(self):
        self.Push_working_history("Trở về lịch sử lỗi sản phẩm hiện tại")
        # Đặt lại các giá trị về mặc định
        table = self.parent.tbErrorProduct
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)
        self.parent.btnDefault_2.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault_2", "#A4A4A4", "15px")
        self.UpdateErrorTimer.start(5000)
        self.UpdateErrorHistory()
        # Đặt lại giá trị của các QDateTimeEdit về ngày hiện tại
        self.current_date = QDate.currentDate()
        one_week_ago = self.current_date.addDays(-6)
        self.parent.dtDayStartHistory1.setDate(one_week_ago)
        self.parent.dtDayFinishHistory1.setDate(self.current_date)

    def ExportHistory(self):
        self.Push_working_history("Xuất file excel lịch sử sản xuất")
        #Xuất từ table tbManufacturingHistory hiện tại ra file .xlsx
        table = self.parent.tbManufacturingHistory
        row_count = table.rowCount()
        column_count = table.columnCount()
        data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = table.item(row, column)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)
        #Tạo file excel
        workbook = Workbook()
        worksheet = workbook.active
        #Tạo tên thêm ngày tháng năm hiện tại
        current_datetime = QDateTime.currentDateTime()
        current_date_str = current_datetime.toString("yyyy-MM-dd-HH-mm-ss")
        filename = f"Manufacturing_History_{current_date_str}.xlsx"
        #Tạo tiêu đề cho file excel
        headers = ["STT", "Tên sản phẩm", "Tổng số sản phẩm", "Tổng số sản phẩm đạt", "Tổng số sản phẩm không đạt", "Tổng số nắp bị lỗi", "Tổng số nắp bị thiếu", "Thời gian bắt đầu sản xuất", "Thời gian kết thúc sản xuất"]
        worksheet.append(headers)
        #Thêm dữ liệu vào file excel
        for row in data:
            worksheet.append(row)
        # Tự động tính độ rộng từng cột
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Lấy chỉ số cột (int)
            column_letter = get_column_letter(column)  # Chuyển sang A, B, C,...

            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            # Cộng thêm 2 để có khoảng trắng
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = alignment
                    cell.border = thin_border

        #Lưu file excel ở H:\APP UNIVERSITY\CODE PYTHON
        workbook.save(f"H:\\APP UNIVERSITY\\CODE PYTHON\\{filename}")
        #Thông báo đã xuất file thành công
        print("Đã xuất file thành công:", filename)

    def ExportErrorHistory(self):
        self.Push_working_history("Xuất file excel lịch sử lỗi sản phẩm")
        #Xuất từ table tbManufacturingHistory hiện tại ra file .xlsx
        table = self.parent.tbErrorProduct
        row_count = table.rowCount()
        column_count = table.columnCount()
        data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = table.item(row, column)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)
        #Tạo file excel
        workbook = Workbook()
        worksheet = workbook.active
        #Tạo tên thêm ngày tháng năm hiện tại
        current_datetime = QDateTime.currentDateTime()
        current_date_str = current_datetime.toString("yyyy-MM-dd-HH-mm-ss")
        filename = f"Error_History_{current_date_str}.xlsx"
        #Tạo tiêu đề cho file excel
        headers = ["STT", "Tên lỗi", "Tên sản phẩm", "Thứ tự khay", "Vị trí lỗi", "Thời gian xảy ra lỗi"]
        worksheet.append(headers)
        #Thêm dữ liệu vào file excel
        for row in data:
            worksheet.append(row)
        # Tự động tính độ rộng từng cột
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Lấy chỉ số cột (int)
            column_letter = get_column_letter(column)  # Chuyển sang A, B, C,...

            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            # Cộng thêm 2 để có khoảng trắng
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = alignment
                    cell.border = thin_border

        #Lưu file excel ở H:\APP UNIVERSITY\CODE PYTHON
        workbook.save(f"H:\\APP UNIVERSITY\\CODE PYTHON\\{filename}")
        #Thông báo đã xuất file thành công
        print("Đã xuất file thành công:", filename)
        
    def Push_working_history(self, action_name: str):
        action_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
        c = self.conn.cursor()
        #Lưu vào database
        c.execute("INSERT INTO WorkingHistory (FullName, Duty, ActionName, ActionTime) VALUES (?, ?, ?, ?)",
                (self.initial.FullName[0], self.initial.Duty[0], action_name, action_time))
        
        self.conn.commit()
            
       


                    
            
        

        
        