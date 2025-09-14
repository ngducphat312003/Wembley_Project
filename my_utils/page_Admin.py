from MainUI1 import Ui_MainWindow
from cryptography.fernet import Fernet
import bcrypt
from PyQt6.QtWidgets import QTableWidgetItem, QAbstractItemView
from PyQt6.QtCore import Qt, QDate, QTimer, QDateTime
from PyQt6 import QtWidgets
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


class AdminBackend():
    def __init__(self, main: Ui_MainWindow, conn, KeyPrimary):
        self.parent = main
        self.conn = conn
        self.KeyPrimary = KeyPrimary
        self.checkbox = None
        self.action_count = 0
        self.FullName = []
        self.Duty = []
        #Hiển thị tên người dùng hiện tại thông qua db
        self.ImportDataFromDB()
        #Setup style cho các bảng
        self.SetupStyleTable(self.parent.tbWorkingHistoryAdmin)
        self.SetupStyleTable(self.parent.tbEmployeeList)
        #Setup style cho các QDateTimeEdit
        self.SetStylePopup(self.parent.dtDayStartHistory1_2)
        self.SetStylePopup(self.parent.dtDayFinishHistory1_2)
        #Khóa ngày chọn
        self.parent.dtDayStartHistory1_2.dateChanged.connect(self.lock_date_range)    
        self.parent.dtDayFinishHistory1_2.dateChanged.connect(self.lock_date_range)
        #Hiển thị danh sách nhân viên hiện tại
        self.ShowEmployeeList()
        #Trang mặc định
        self.parent.stackedWidget_3.setCurrentWidget(self.parent.page_EmployeeWorking)
        #Tô màu background cho nút btnEmployeeWorking và btnEmployeeList
        self.SetStyleSheetForbtn("btnEmployeeWorking", "#33D909")
        self.SetStyleSheetForbtn("btnEmployeeList", "#A4A4A4")
        self.parent.btnDefault_3.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault_3", "#A4A4A4")
        #Reset nút btnEmployeeCreation khi chuyển trang
        self.parent.stackedWidget.currentChanged.connect(lambda: [self.ModifyEmployee(False)])
        #Nút nhấn tạo mới nhân viên
        self.CreateEmployee()
        #Nút nhấn thay đổi trang EmployeeWorking và EmployeeList
        self.parent.btnEmployeeWorking.clicked.connect((lambda: [self.parent.stackedWidget_3.setCurrentWidget(self.parent.page_EmployeeWorking), self.WorkingHistoryAdmin()]))
        self.parent.btnEmployeeList.clicked.connect((lambda: [self.parent.stackedWidget_3.setCurrentWidget(self.parent.page_EmployeeList), self.EmployeeListAdmin()]))
        #Nhấn nút lưu thông tin nhân viên vừa tạo
        self.parent.btnSave1.clicked.connect((lambda: [self.CreatedEmployee(), self.ModifyEmployee(False)]))
        #Nút nhấn xóa nhân viên
        self.parent.btnDelete.clicked.connect(self.DeleteEmployee)
        #Nút nhấn sửa thông tin nhân viên
        self.parent.btnModify.clicked.connect((lambda: [self.ModifyEmployeeInfo(), self.ModifyEmployee1()]))

        #tbWorkingHistoryAdmin
        self.Screening_Current_Date()
        #Câp nhật lại danh sách nhân viên
        self.UpdateHandleHistory()
        #Chạy timer để cập nhật lại danh sách nhân viên
        self.UpdateHandleTimer = QTimer()
        self.UpdateHandleTimer.timeout.connect(self.UpdateHandleHistoryTimer)
        self.UpdateHandleTimer.start(5000)
        
        #Nút nhấn lọc lịch sử làm việc
        self.parent.btnFilter2_2.clicked.connect(self.FilterHistory)
        #Nút nhấn đặt lại lịch sử làm việc
        self.parent.btnDefault_3.clicked.connect(self.ResetHistory)
        #Nút nhấn xuất dữ liệu lịch sử làm việc
        self.parent.btnExport_3.clicked.connect(self.ExportHistory)
        
        self.parent.cboxModifyPara_2.setChecked(False)
        self.parent.cboxModifyPara_3.setChecked(False)
        self.parent.cboxModifyPara_4.setChecked(False)
        self.parent.cboxModifyPara_5.setChecked(False)
        self.parent.cboxModifyPara_2.setEnabled(False)
        self.parent.cboxModifyPara_3.setEnabled(False)
        self.parent.cboxModifyPara_4.setEnabled(False)
        self.parent.cboxModifyPara_5.setEnabled(False)
       
    def ImportDataFromDB(self):
        c = self.conn.cursor()
        #Lấy rowid từ Account
        RowId = c.execute("SELECT rowid FROM Account WHERE UserName = ?", (self.KeyPrimary,)).fetchall()
        #Hiển thị FullName từ Account
        self.FullName = c.execute("SELECT FullName FROM Account WHERE rowid = ?", RowId[0]).fetchone()
        self.parent.lbFullName.setText(self.FullName[0])
        #Hiển thị Duty từ Account
        self.Duty = c.execute("SELECT Duty FROM Account WHERE rowid = ?", RowId[0]).fetchone()
        self.parent.lbDuty.setText(self.Duty[0])
    #Bảng WorkingHistory
    def WorkingHistoryAdmin(self):
        self.Push_working_history("Truy cập lịch sử làm việc của nhân viên")
        #Setup độ rộng cho các cột trong bảng tbWorkingHistoryAdmin
        #Tô màu background cho nút btnEmployeeWorking và btnEmployeeList
        self.SetStyleSheetForbtn("btnEmployeeWorking", "#33D909")
        self.SetStyleSheetForbtn("btnEmployeeList", "#A4A4A4")

    def Screening_Current_Date(self):
        self.current_date = QDate.currentDate()
        one_week_ago = self.current_date.addDays(-6)
        #Hiển thị trên QDateTimeEdit
        self.parent.dtDayStartHistory1_2.setDate(one_week_ago)
        self.parent.dtDayFinishHistory1_2.setDate(self.current_date)

    def UpdateHandleHistory(self):
        current_date = QDate.currentDate()
        c = self.conn.cursor()
        query = """
            SELECT * FROM WorkingHistory
            WHERE DATE(ActionTime) = ?
        """
        c.execute(query, (current_date.toString("yyyy-MM-dd"),))
        actions = c.fetchall()
        table = self.parent.tbWorkingHistoryAdmin
        current_row_count = table.rowCount()  # Lấy số dòng hiện tại của bảng
        table.verticalHeader().setVisible(False)
        rows = actions[::-1]
        self.action_count = len(rows)
        for row_idx, row_data in enumerate(rows):
            table.insertRow(current_row_count + row_idx)
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
             # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 4:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

    def UpdateHandleHistoryTimer(self):
        #Lấy ngày hiện tại
        current_date = QDate.currentDate()
        #Thực hiện truy vấn SQL để lấy ngày muộn nhất trong StartTime
        c = self.conn.cursor()
        query = """
            SELECT * FROM WorkingHistory
            WHERE DATE(ActionTime) = ?
        """
        c.execute(query, (current_date.toString("yyyy-MM-dd"),))
        actions = c.fetchall()
        rows = actions[::-1]
        if len(rows) != self.action_count:
            table = self.parent.tbWorkingHistoryAdmin
            table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
            current_row_count = table.rowCount()  # Lấy số dòng hiện tại của bảng
            for row_idx, row_data in enumerate(rows):
                table.insertRow(current_row_count + row_idx)
                # Thêm cột số thứ tự vào cột đầu tiên (STT)
                item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
                # Thêm các giá trị khác vào các cột còn lại
                for col_idx, value in enumerate(row_data):
                    if col_idx < 5:  # Tránh việc truy cập ngoài phạm vi
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự
    #Bảng EmployeeList
    def EmployeeListAdmin(self):
        self.Push_working_history("Truy cập danh sách nhân viên")
        #Tô màu background cho nút btnEmployeeWorking và btnEmployeeList
        self.SetStyleSheetForbtn("btnEmployeeWorking", "#A4A4A4")
        self.SetStyleSheetForbtn("btnEmployeeList", "#33D909")
    #Hiển thị danh sách nhân viên hiện tại
    def ShowEmployeeList(self):
        #Lấy danh sách nhân viên từ database
        self.cursor = self.conn.cursor()
        self.cursor.execute("SELECT Fullname, Duty, Role FROM Account")
        info = self.cursor.fetchall()
        table = self.parent.tbEmployeeList
        #Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
        current_row_count = table.rowCount()  # Lấy số dòng hiện tại của bảng
        table.verticalHeader().setVisible(False)
        #Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(info):
            table.insertRow(current_row_count + row_idx)  # Thêm hàng mới vào cuối bảng
            
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(current_row_count + row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(current_row_count + row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 5:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(current_row_count + row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự
                #Thêm Qcheckbox vào các cột cuối cùng của bảng QtWidgets.QCheckBox
            self.checkbox = QtWidgets.QCheckBox()
            self.checkbox.setStyleSheet("border: None;\n"
                        "border-radius: 5px;\n"
                        "font-family: Inter, sans-serif;\n"
                        "color: black;\n"
                        "margin-left: 80%;\n"
                        "\n"
                        "")
            self.checkbox.setTristate(False) 
            table.setCellWidget(current_row_count + row_idx, 4, self.checkbox)
            self.checkbox.stateChanged.connect(lambda state, r=row_idx: self.checkbox_changed(state, r))

    def checkbox_changed(self, state, row):
        if state == 2:
            #Tắt checkbox của cboxModifyPara_2, cboxModifyPara_3, cboxModifyPara_4, cboxModifyPara_5
            self.parent.cboxModifyPara_2.setChecked(False)
            self.parent.cboxModifyPara_3.setChecked(False)
            self.parent.cboxModifyPara_4.setChecked(False)
            self.parent.cboxModifyPara_5.setChecked(False)
            self.parent.cboxModifyPara_2.setEnabled(False)
            self.parent.cboxModifyPara_3.setEnabled(True)
            self.parent.cboxModifyPara_4.setEnabled(True)
            self.parent.cboxModifyPara_5.setEnabled(True)
            #Lấy thông tin từ QLineEdit và QComboBox
            values = []
            for col in range(self.parent.tbEmployeeList.columnCount() - 1):  # Bỏ cột chứa checkbox
                item = self.parent.tbEmployeeList.item(row, col)
                values.append(item.text() if item else '')
            #Tìm kiếm hàng chứa values[1] trong database
            self.cursor = self.conn.cursor()
            self.cursor.execute("SELECT rowid FROM Account WHERE FullName = ?", (values[1],))
            row_db = self.cursor.fetchone()
            if row_db is None:
                print("Không tìm thấy nhân viên trong database1")
                return
            #Lấy thông tin nhân viên từ database
            self.cursor.execute("SELECT UserName, Fullname, Duty, Role FROM Account WHERE rowid = ?", (row_db[0],))
            info = self.cursor.fetchone()
            #Điền thông tin vào các QLineEdit và QComboBox
            self.parent.leditFullNameEmployee.setText(info[1])
            self.parent.cbDuty1.setCurrentText(info[2])
            self.parent.leditUsername.setText(info[0])
            self.parent.leditPassword.setText("")
            #Điền thông tin vào các QCheckBox
            self.parent.cboxModifyInfor.setChecked("1" in info[3])
            self.parent.cboxModifyPara.setChecked("2" in info[3])
            self.parent.cboxModifyInfor_3.setChecked("3" in info[3])
            self.parent.cboxModifyInfor_4.setChecked("4" in info[3])
            self.parent.cboxModifyInfor_5.setChecked("5" in info[3])
            #Kích hoạt nút btnDelete và btnModify
            self.parent.btnDelete.setEnabled(True)
            self.SetStyleSheetForbtn("btnDelete", "#DF2026")
            self.parent.btnModify.setEnabled(True)
            self.SetStyleSheetForbtn("btnModify", "#F68013")
            #Kích hoạt trong trường QLineEdit và QComboBox
            self.parent.cboxModifyInfor.setEnabled(True)
            self.parent.cboxModifyPara.setEnabled(True)
            self.parent.cboxModifyInfor_3.setEnabled(True)
            self.parent.cboxModifyInfor_4.setEnabled(True)
            self.parent.cboxModifyInfor_5.setEnabled(True)
            # Disable các checkbox khác
            for r in range(self.parent.tbEmployeeList.rowCount()):
                if r != row:
                    checkbox = self.parent.tbEmployeeList.cellWidget(r, self.parent.tbEmployeeList.columnCount() - 1)
                    if checkbox:
                        checkbox.setEnabled(False)
        elif state == 0:
            for r in range(self.parent.tbEmployeeList.rowCount()):
                checkbox = self.parent.tbEmployeeList.cellWidget(r, self.parent.tbEmployeeList.columnCount() - 1)
                if checkbox:
                    checkbox.setEnabled(True)
            # Reset lại các QLineEdit và QComboBox, QCheckBox
            self.Content_Reset()
            self.parent.cboxModifyPara_2.setChecked(False)
            self.parent.cboxModifyPara_3.setChecked(False)
            self.parent.cboxModifyPara_4.setChecked(False)
            self.parent.cboxModifyPara_5.setChecked(False)
            self.parent.cboxModifyPara_2.setEnabled(False)
            self.parent.cboxModifyPara_3.setEnabled(False)
            self.parent.cboxModifyPara_4.setEnabled(False)
            self.parent.cboxModifyPara_5.setEnabled(False)
            # Kích hoạt nút btnDelete và btnModify
            self.parent.btnDelete.setEnabled(False) 
            self.SetStyleSheetForbtn("btnDelete", "#A4A4A4")
            self.parent.btnModify.setEnabled(False)
            self.SetStyleSheetForbtn("btnModify", "#A4A4A4")
            # Kích hoạt nút btnSave1
            self.parent.btnSave1.setEnabled(False)
            self.SetStyleSheetForbtn("btnSave1", "#A4A4A4")

    #Thêm style cho nút 
    def SetStyleSheetForbtn(self, btn, background_color):
        #Style cho nút btnEmployeeWorking
        button = getattr(self.parent, btn)
        button.setStyleSheet(f"""
                QPushButton#{btn} {{
                    border-radius: 15px;
                    border-color: white;
                    background-color: {background_color};  /* Màu nền mới */
                    color: white;
                    text-align: center;
                    font-family: Inter, sans-serif;
                }}

                QPushButton#{btn}:hover {{
                    background-color: #FFA500;  /* Màu nền khi hover */
                }}

                QPushButton#{btn}:pressed {{
                    padding-left: 5px;
                    padding-top: 5px;
                }}
                """)                           
    #Tạo mới nhân viên
    def CreateEmployee(self):
        #Tắt khả năng điều chỉnh QLineEdit và QComboBox, QcheckBox
        self.ModifyEmployee(False)
        self.parent.btnSave1.setEnabled(False)
        #Nút nhấn tạo mới nhân viên
        self.parent.btnEmployeeCreation.clicked.connect((lambda: [self.ModifyEmployee(command = True), self.parent.btnSave1.setEnabled(True),self.SetStyleSheetForbtn("btnSave1", "#0055FF"), self.Content_Reset(), self.parent.stackedWidget_3.setCurrentWidget(self.parent.page_EmployeeList), self.SetStyleSheetForbtn("btnEmployeeWorking", "#A4A4A4"), self.parent.btnEmployeeList.setEnabled(True), self.SetStyleSheetForbtn("btnEmployeeList", "#33D909"), self.Push_working_history("Tạo mới nhân viên")]))
    #Cho phép điều chỉnh
    def ModifyEmployee(self,command):
        #Nhấn check cho qcheckbox
        self.parent.cboxModifyPara_2.setChecked(command)
        self.parent.cboxModifyPara_3.setChecked(command)
        self.parent.cboxModifyPara_4.setChecked(command)
        self.parent.cboxModifyPara_5.setChecked(command)
        if self.parent.cboxModifyPara_2.isChecked():
            self.parent.leditFullNameEmployee.setEnabled(True)
        else:
            self.parent.cboxModifyPara_2.setEnabled(False)
        if self.parent.cboxModifyPara_3.isChecked():
            self.parent.cbDuty1.setEnabled(True)
        else:
            self.parent.cboxModifyPara_3.setEnabled(False)
        if self.parent.cboxModifyPara_4.isChecked():
            self.parent.leditUsername.setEnabled(True)
        else:
            self.parent.cboxModifyPara_4.setEnabled(False)
        if self.parent.cboxModifyPara_5.isChecked():
            self.parent.leditPassword.setEnabled(True)
        else:
            self.parent.cboxModifyPara_5.setEnabled(False)
        
        self.parent.cboxModifyInfor.setEnabled(command)
        self.parent.cboxModifyPara.setEnabled(command)
        self.parent.cboxModifyInfor_3.setEnabled(command)
        self.parent.cboxModifyInfor_4.setEnabled(command)
        self.parent.cboxModifyInfor_5.setEnabled(command)
        #Tắt nút khả năng nhấn của btnDelete
        self.parent.btnDelete.setEnabled(False)
        self.SetStyleSheetForbtn("btnDelete", "#A4A4A4")
        self.parent.btnModify.setEnabled(False)
        self.SetStyleSheetForbtn("btnModify", "#A4A4A4")
        self.parent.btnSave1.setEnabled(False)
        self.SetStyleSheetForbtn("btnSave1", "#A4A4A4")

    def ModifyEmployee1(self):
        self.ModifyEmployee(False)
        #Reset nội dung của các QLineEdit và QComboBox, QCheckBox
        self.Content_Reset()
        #Reset Qcheckbox của bảng EmployeeList
        for r in range(self.parent.tbEmployeeList.rowCount()):
            checkbox = self.parent.tbEmployeeList.cellWidget(r, self.parent.tbEmployeeList.columnCount() - 1)
            if checkbox:
                checkbox.setChecked(False)
                checkbox.setEnabled(True)
    #Lưu thông tin nhân viên vừa tạo
    def CreatedEmployee(self):
        self.Push_working_history("Lưu thông tin nhân viên vừa tạo")
        self.cursor = self.conn.cursor()        
        #Lấy thông tin từ các QLineEdit và QComboBox
        full_name = self.parent.leditFullNameEmployee.text()
        duty = self.parent.cbDuty1.currentText()
        username = self.parent.leditUsername.text()
        password = self.parent.leditPassword.text()
        #Mã hóa mật khẩu
        hashed_password = self.create_password_hash(password)
        #Lấy thông tin từ các QCheckBox
        checkboxes = [
                        (self.parent.cboxModifyInfor, "1"),
                        (self.parent.cboxModifyPara, "2"),
                        (self.parent.cboxModifyInfor_3, "3"),
                        (self.parent.cboxModifyInfor_4, "4"),
                        (self.parent.cboxModifyInfor_5, "5"),
                    ]

        roles = "".join(num for cb, num in checkboxes if cb.isChecked())
        #Thông tin PLC
        PLCHost = "192.168.100.14"
        #Mã hóa địa chỉ IP
        key = Fernet.generate_key() 
        print("key", key)
        encrypted_ip = self.encrypt_ip(PLCHost, key)
        PLCPort = 4095
        PLCType = "Q"
        #Lưu thông tin vào database
        self.cursor.execute("INSERT INTO Account (Username, Password, FullName, Duty, Role, PLCHost, PLCPort, PLCType, key) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                            (username, hashed_password, full_name, duty, roles, encrypted_ip, PLCPort, PLCType, key))
        self.conn.commit()
        print("Đã lưu thông tin nhân viên vào database")
        #Cập nhật lại danh sách nhân viên
        self.parent.tbEmployeeList.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
        self.ShowEmployeeList()
    #Mã hóa mật khẩu
    def create_password_hash(self, password: str):
        salt = bcrypt.gensalt()
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt)
        return hashed_password
    # Mã hóa và giải mã IP với Fernet
    def encrypt_ip(self, ip_address: str, key: bytes) -> str:
        fernet = Fernet(key)
        encrypted_ip = fernet.encrypt(ip_address.encode()).decode()
        return encrypted_ip
    #Reset nội dung của các QLineEdit và QComboBox
    def Content_Reset(self):
        #Reset nội dung của các QLineEdit và QComboBox
        self.parent.leditFullNameEmployee.setText("")
        self.parent.cbDuty1.setCurrentIndex(0)
        self.parent.leditUsername.setText("")
        self.parent.leditPassword.setText("")
        #Reset nội dung của các QCheckBox
        self.parent.cboxModifyInfor.setChecked(False)
        self.parent.cboxModifyPara.setChecked(False)
        self.parent.cboxModifyInfor_3.setChecked(False)
        self.parent.cboxModifyInfor_4.setChecked(False)
        self.parent.cboxModifyInfor_5.setChecked(False)
    #Xóa thông tin nhân viên
    def DeleteEmployee(self):
        self.Push_working_history("Xóa thông tin nhân viên")
        self.cursor = self.conn.cursor()
        #Lấy thông tin từ QLineEdit
        fullname = self.parent.leditFullNameEmployee.text()
        #Lấy hàng chứa username trong database
        row = self.cursor.execute("SELECT rowid FROM Account WHERE FullName = ?", (fullname,))
        row = row.fetchone()
        if row is None:
            print("Không tìm thấy nhân viên trong database2")
            return
        #Xóa thông tin nhân viên trong database
        self.cursor.execute("DELETE FROM Account WHERE rowid = ?", (row[0],))
        self.conn.commit()
        print("Đã xóa thông tin nhân viên trong database")
        #Reset nội dung của các QLineEdit và QComboBox
        self.ModifyEmployee1()
        #Cập nhật lại danh sách nhân viên
        self.parent.tbEmployeeList.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng
        self.ShowEmployeeList()
    #Sửa thông tin nhân viên
    def ModifyEmployeeInfo(self):
        self.Push_working_history("Sửa thông tin nhân viên")
        self.cursor = self.conn.cursor()
        #Lấy thông tin từ QLineEdit và QComboBox
        full_name = self.parent.leditFullNameEmployee.text()
        duty = self.parent.cbDuty1.currentText()
        username = self.parent.leditUsername.text()
        password = self.parent.leditPassword.text()
        #Mã hóa mật khẩu
        hashed_password = self.create_password_hash(password)
        #Lấy thông tin từ các QCheckBox
        checkboxes = [
                        (self.parent.cboxModifyInfor, "1"),
                        (self.parent.cboxModifyPara, "2"),
                        (self.parent.cboxModifyInfor_3, "3"),
                        (self.parent.cboxModifyInfor_4, "4"),
                        (self.parent.cboxModifyInfor_5, "5"),
                    ]

        roles = "".join(num for cb, num in checkboxes if cb.isChecked())
        #Lấy hàng chứa full_name trong database
        row = self.cursor.execute("SELECT rowid FROM Account WHERE FullName = ?", (full_name,))
        row = row.fetchone()
        if row is None:
            print("Không tìm thấy nhân viên trong database3")
            return
        #Sửa thông tin nhân viên trong database
        
        #Kiểm tra các QCheckBox
        if self.parent.cboxModifyPara_2.isChecked():
            self.parent.leditFullNameEmployee.setEnabled(True)
            self.cursor.execute("UPDATE Account SET FullName = ? WHERE rowid = ?", (full_name, row[0]))
        else:
            self.parent.leditFullNameEmployee.setEnabled(False)
        if self.parent.cboxModifyPara_3.isChecked():
            self.parent.cbDuty1.setEnabled(True)
            self.cursor.execute("UPDATE Account SET Duty = ? WHERE rowid = ?", (duty, row[0]))
        else:
            self.parent.cbDuty1.setEnabled(False)
        if self.parent.cboxModifyPara_4.isChecked():
            self.parent.leditUsername.setEnabled(True)
            self.cursor.execute("UPDATE Account SET UserName = ? WHERE rowid = ?", (username, row[0]))
        else:
            self.parent.leditUsername.setEnabled(False)
        if self.parent.cboxModifyPara_5.isChecked():
            self.parent.leditPassword.setEnabled(True)
            self.cursor.execute("UPDATE Account SET PassWord = ? WHERE rowid = ?", (hashed_password ,row[0]))
        else:
            self.parent.leditPassword.setEnabled(False)

        self.cursor.execute("UPDATE Account SET Role = ? WHERE rowid = ?", (roles ,row[0]))

        self.conn.commit()
        self.ShowEmployeeList()
            
    def SetupStyleTable(self, table):
        scrollbar_v = self.scrollbar("vertical", "ffffff", "0055FF")
        scrollbar_h = self.scrollbar("horizontal", "ffffff", "0055FF")
        table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        table.setShowGrid(False)
        style = """
                QTableWidget {
                    background-color: #ffffff;  /* Màu nền bảng */
                    border-radius: 0px;  /* Bo tròn góc */
                    color: #0055FF;  /* Màu chữ */
                    font-family: Inter, sans-serif;
                    font-size: 14px;
                    font-weight: bold;
                    border: 1px solid #0055FF;
                }
                QTableWidget::item {
                    border: 1px solid #0055FF;  /* Viền đỏ cho tất cả ô */
                    font-family: Inter, sans-serif;
                    font-size: 14px;
                    font-weight: bold;
                }
                QHeaderView::section {
                background-color: #ffffff;
                border: 1px solid #0055FF;  /* Ẩn viền */
                font-weight: bold;
                font-family: Inter, sans-serif;
                font-size: 16px;
                }
                """
        full_style = style + scrollbar_v + scrollbar_h

        table.setStyleSheet(full_style)

    def SetStylePopup(self, datetime_edit):
        datetime_edit.setCalendarPopup(True)
        calendar = datetime_edit.calendarWidget()
        calendar.setStyleSheet("""
            QCalendarWidget QWidget { 
                background-color: #FFFFFF;
                border: None; 
                border-radius: 0px;  
            }

            QCalendarWidget QToolButton {
                background-color: white;
                color: #1b731b;
                font: bold 14px;
                border: none;
                margin: 5px;
                padding: 5px;
                border-radius: 0px;
            }

            QCalendarWidget QMenu {
                background-color: white;
                color: black;
                border-radius: 0px;
            }

            QCalendarWidget QSpinBox { 
                width: 60px; 
                font-size: 14px;
                background-color: black; 
            }

            QCalendarWidget QAbstractItemView {
                font-size: 14px;
                background-color: white;
                selection-background-color: #1b731b;
                selection-color: white;
                gridline-color: #FFFFFF;
            }

            QCalendarWidget QAbstractItemView:enabled {
                color: #333;
            }

            QCalendarWidget QAbstractItemView:disabled {
                color: #999;
            }
        """)
        
    def scrollbar(self, direction, base_color, color):
        style = f"""
        QScrollBar:{direction} {{
            border: none; 
            background-color: #{base_color}; 
            {'width' if direction == 'vertical' else 'height'}: 10px; 
            margin: 0px 0px 0px 0px; 
        }}
        
        QScrollBar::handle:{direction} {{
            background-color: #{color}; 
            border-radius: 5px; 
        }}
        
        QScrollBar::handle:{direction}:hover {{
            background-color: #4D9EFF; 
        }}
        
        QScrollBar::add-line:{direction}, QScrollBar::sub-line:{direction} {{
            border: none; 
            background: none; 
        }}
        
        QScrollBar::add-page:{direction}, QScrollBar::sub-page:{direction} {{
            background: none;
        }}
        """
        return style
        
    def lock_date_range(self):
        # Lấy ngày hiện tại
        current_date = QDate.currentDate()
        # Thực hiện truy vấn SQL để lấy ngày muộn nhất trong StartTime
        c = self.conn.cursor()
        query = """
            SELECT rowid, ActionTime FROM WorkingHistory
            ORDER BY ActionTime ASC LIMIT 1
        """
        c.execute(query)
        result = c.fetchone()
        # print("result", result)
        if result:
            # Lấy ngày muộn nhất từ StartTime (cột thứ 2 trong kết quả)
            latest_start_date = QDate.fromString(result[1].split()[0], "yyyy-MM-dd")

        # Khóa dtDayStartHistory (không cho phép chọn từ ngày hôm nay trở đi)
        self.parent.dtDayStartHistory1_2.setDateRange(latest_start_date, current_date)
        start_date = self.parent.dtDayStartHistory1_2.date()
        # Khóa dtDayFinishHistory (không cho phép chọn từ ngày hôm sau trở đi)
        self.parent.dtDayFinishHistory1_2.setDateRange(start_date, current_date)

        self.parent.dtDayStartHistory1_2.setDateRange(latest_start_date, current_date)
        start_date1 = self.parent.dtDayStartHistory1_2.date()
        # Khóa dtDayFinishHistory (không cho phép chọn từ ngày hôm sau trở đi)
        self.parent.dtDayFinishHistory1_2.setDateRange(start_date1, current_date)

    def FilterHistory(self):
        self.Push_working_history("Lọc lịch sử làm việc")
        self.parent.btnDefault_3.setEnabled(True)
        self.SetStyleSheetForbtn("btnDefault_3", "#33D909")
        
        self.UpdateHandleTimer.stop()
        start_date = self.parent.dtDayStartHistory1_2.date()
        end_date = self.parent.dtDayFinishHistory1_2.date()

        c = self.conn.cursor()
        query = """
            SELECT * FROM WorkingHistory
            WHERE DATE(ActionTime) BETWEEN ? AND ?
        """
        c.execute(query, (start_date.toString("yyyy-MM-dd"), end_date.toString("yyyy-MM-dd")))
        info = c.fetchall()
        table = self.parent.tbWorkingHistoryAdmin
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)  # Xóa tất cả các hàng hiện tại trong bảng

        rows = info[::-1]
        # Thêm dữ liệu vào bảng
        for row_idx, row_data in enumerate(rows):
            table.insertRow(row_idx)  # Thêm hàng mới vào cuối bảng
            
            # Thêm cột số thứ tự vào cột đầu tiên (STT)
            item = QTableWidgetItem(str(row_idx + 1))  # Tính số thứ tự
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row_idx, 0, item)  # Đặt số thứ tự vào cột đầu tiên
            
            # Thêm các giá trị khác vào các cột còn lại
            for col_idx, value in enumerate(row_data):
                if col_idx < 4:  # Tránh việc truy cập ngoài phạm vi
                    item = QTableWidgetItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    table.setItem(row_idx, col_idx + 1, item)  # Đặt vào các cột sau cột số thứ tự

    def ResetHistory(self):
        self.Push_working_history("Trở về lịch sử làm việc hiện tại")
        # Đặt lại các giá trị về mặc định
        table = self.parent.tbWorkingHistoryAdmin
        # Xóa tất cả các hàng hiện tại trong bảng trước khi thêm dữ liệu mới
        table.setRowCount(0)
        self.parent.btnDefault_3.setEnabled(False)
        self.SetStyleSheetForbtn("btnDefault_3", "#A4A4A4")
        self.UpdateHandleTimer.start(5000)
        self.UpdateHandleHistory()
        # Đặt lại giá trị của các QDateTimeEdit về ngày hiện tại
        self.current_date = QDate.currentDate()
        one_week_ago = self.current_date.addDays(-6)
        self.parent.dtDayStartHistory1_2.setDate(one_week_ago)
        self.parent.dtDayFinishHistory1_2.setDate(self.current_date)    

    def ExportHistory(self):
        self.Push_working_history("Xuất file excel lịch sử làm việc")
        #Xuất từ table tbWorkingHistoryAdmin hiện tại ra file .xlsx
        table = self.parent.tbWorkingHistoryAdmin
        row_count = table.rowCount()
        column_count = table.columnCount()
        data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = table.item(row, column)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)
        #Tạo file excel
        workbook = Workbook()
        worksheet = workbook.active
        #Tạo tên thêm ngày tháng năm hiện tại
        current_datetime = QDateTime.currentDateTime()
        current_date_str = current_datetime.toString("yyyy-MM-dd-HH-mm-ss")
        filename = f"Working_History_{current_date_str}.xlsx"
        #Tạo tiêu đề cho file excel
        headers = ["STT", "Tên nhân viên", "Chức vụ", "Tên thao tác", "Thời gian thực hiện"]
        worksheet.append(headers)
        #Thêm dữ liệu vào file excel
        for row in data:
            worksheet.append(row)
        # Tự động tính độ rộng từng cột
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Lấy chỉ số cột (int)
            column_letter = get_column_letter(column)  # Chuyển sang A, B, C,...

            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            # Cộng thêm 2 để có khoảng trắng
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = alignment
                    cell.border = thin_border

        #Lưu file excel ở H:\APP UNIVERSITY\CODE PYTHON
        workbook.save(f"H:\\APP UNIVERSITY\\CODE PYTHON\\{filename}")
        #Thông báo đã xuất file thành công
        print("Đã xuất file thành công:", filename)
    
    def Push_working_history(self, action_name: str):
        action_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
        c = self.conn.cursor()
        #Lưu vào database
        c.execute("INSERT INTO WorkingHistory (FullName, Duty, ActionName, ActionTime) VALUES (?, ?, ?, ?)",
                (self.FullName[0], self.Duty[0], action_name, action_time))
        
        self.conn.commit()

    def connect_checkbox_to_widget(self, checkbox: QtWidgets.QCheckBox, widget: QtWidgets.QWidget):
        def toggle_widget(state):
            print("state", state)
            widget.setEnabled(state == Qt.CheckState.Checked)
        checkbox.stateChanged.connect(toggle_widget)    
